# モジュール読み込み
import calendar, datetime, openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 必要なリストとディクショナリの用意
workerlist = {}                                         # 従業員全員
workabledays = {}                                       # 従業員が勤務可能な日
workingdays = {}
weekdays = []                                           # 平日のリスト
satlist = []                                            # 土曜日のリスト
holydaylist = []                                        # 日祝のリスト
daysInTheMonth = []                                     # 月の日にちのリスト
pubholys = {4:[30] ,5:[3, 4, 5], 7:[16], 8:[11], 9:[17, 24], 10:[8], 11:[3, 23], 12:[24]}
partWorkers = []    # 土日要員のリスト
fullWorkers = []    # 平日要員のリスト
dates =[]           # 曜日表示用のリスト
datesDic = {0:"月",1:"火",2:"水",3:"木",4:"金",5:"土",6:"日"}
shift = {}          # 最終的なシフトを格納するディクショナリ
noWorkingDays = {4:8,5:6,6:10,7:8,8:12,9:9,10:8,11:11,12:9} # 休刊日のディクショナリ

# 月に関する情報の取得

nextmonth = int(datetime.date.today().month + 1)        # シフトを作る月を取得
print(f"{nextmonth}月分のシフトを作成します")

if nextmonth in pubholys:
    holydaylist.extend(pubholys[nextmonth])     # 祝日をリストに追加

print(calendar.month(2018, nextmonth))

(a, b) = calendar.monthrange(2018, nextmonth)           # 月の最初の曜日と日数を取得

for day in range(1, b+1):
    daysInTheMonth.append(day)                          # 月の日にちをリストに突っ込む

if (0 <= a) and (a <= 5):                               # 月の最初の曜日でパターン分け
    sat = 6 - a
    sun = sat + 1

elif a == 6:
    sat = 7
    sun = 1

thismonth = calendar.Calendar(firstweekday=a)
dateslist = list(thismonth.iterweekdays()) *5   # 曜日を表す数字をdateslistに月の日数分格納


for datenum in dateslist:           # 曜日を表す数字から曜日名に変換し、datesに格納
    if len(dates) <= b-1:
        dates.append(datesDic[datenum])


for i in range(0, 6):                                   # 土曜と日祝をリストに突っ込む
    if (sat + 7 * i) <= b and (sat + 7 * i) not in satlist:
        satlist.append(sat + 7 * i)
        satlist = list(set(satlist) - set(holydaylist))
        satlist.sort()
    if (sun + 7 * i) <= b:
        holydaylist.append(sun + 7 * i)
        holydaylist = list(set(holydaylist))
        holydaylist.sort()

weekdays = list(set(daysInTheMonth) - set(satlist) - set(holydaylist)) #平日のリスト
print(f"{nextmonth}月の祝日は{pubholys[nextmonth]}日、休刊日は{noWorkingDays[nextmonth]}です")

# 従業員に関する情報の取得
while True:                                                 # 従業員の名前と属性を取得
    pname = input("従業員の名前： ")
    if pname in workerlist:
        print("その人はすでに登録されてるよ")
        continue
    ptype = input("従業員の働き方（フルタイム/土日）")
    if (ptype != "フルタイム") and (ptype != "土日"):
        print("フルタイムか土日で入力してね")
        continue

    dayoffneeds = input(f"{pname}の休み希望を入力してください\n例：3,12,13,19,20\n>> ")
    dayofflist = dayoffneeds.split(',') # 休みのリスト

    if ptype == "土日":
        dayofflist = list(set(dayofflist) | set(weekdays))
        partWorkers.append(pname)
    else:
        fullWorkers.append(pname)
    try:
        dayofflist = list(map(int, dayofflist)) # 休みのリストの中身をint型に変換
    except ValueError:
        pass
    handle = ""
    for i in dayofflist:
        if i != "":
            if i not in daysInTheMonth:
                print(f"{i}日は{nextmonth}月に存在しないよー")
                handle = "X"
    if handle == "X":
        continue
    workerinfo = {}
    workerinfo["属性"] = ptype
    workerinfo["休み希望"] = dayofflist
    workerlist[pname] = workerinfo

    workabledays[pname] = list(set(daysInTheMonth) - set(dayofflist))
    if noWorkingDays[nextmonth] in workabledays[pname]:
        workabledays[pname].remove(noWorkingDays[nextmonth])
    print(f"現在登録されている従業員は{workerlist}")
    loopdecider = input('他にも従業員おる？(yes/no) ')
    if loopdecider == "no":
        break

daysforexcel = [f"{nextmonth}月"] + daysInTheMonth


# exelシート作成
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = f"{nextmonth}月"

wb.create_sheet(index=1, title="workerlist")

sh = wb["workerlist"]
sh.merge_cells('A1:C1')
sh['A1'].font = Font(name="sans-serif",size=15, bold=True)
sh['A1']=("勤務可能日")

sh.append(daysforexcel)

# 土日の背景色設定
satcell = PatternFill(
    patternType = 'solid',
    start_color='ff00ff00',
    end_color='ff0000ff')
suncell = PatternFill(
    patternType = 'solid',
    start_color='ffff0000',
    end_color='ffff0000')

for i in range(2,b+2):
    selected = sh.cell(row=2, column=i)
    if selected.value in satlist:
        selected.fill = satcell
    elif selected.value in holydaylist:
        selected.fill = suncell
workers = list(workerlist.keys())

# 従業員の勤務可能日をシートに反映
for j in range(len(workers)):
    tictac = []
    for days in daysInTheMonth:
        if days in workabledays[workers[j]]:
            tictac.append("○")
        else:
            tictac.append("×")
    workableexcel = [workers[j]] + tictac
    sh.append(workableexcel)
wb.save('シフト表.xlsx')

checklist = weekdays + weekdays + satlist + satlist + holydaylist
checklist.sort()

# 従業員の出勤日を確定
for worker in sh['A']:
    memo = []
    if worker.value in partWorkers:     # 土日要員の出勤日
        for sat in satlist:
            if sh.cell(row=worker.row, column=sat+1).value == "○":
                memo.append(sat)
                checklist.remove(sat)

        holyworking= {}
        count = 0
        for hol in holydaylist:

            try:
                if (sh.cell(row=worker.row, column=hol+1).value =="○") and (sh.cell(row=worker.row+1, column=hol+1).value !="○") and (hol in checklist):
                    memo.append(hol)
                    checklist.remove(hol)
                elif (sh.cell(row=worker.row, column=hol+1).value =="○") and (sh.cell(row=worker.row+1, column=hol+1).value =="○"):
                    if (holydaylist[count-1] not in memo) and (hol in checklist):
                        memo.append(hol)

                        checklist.remove(hol)
            except IndexError:
                if (sh.cell(row=worker.row, column=hol+1).value =="○") and (hol in checklist):
                    memo.append(hol)
                    checklist.remove(hol)
            count += 1
    elif worker.value in fullWorkers:   # フルタイムの勤務日
        for day in weekdays:
            if sh.cell(row=worker.row, column=day+1).value =="○":
                memo.append(day)
                checklist.remove(day)

    workingdays[worker.value] = memo
workingdays.pop("勤務可能日")
workingdays.pop(f"{nextmonth}月")
try:
    for i in holydaylist:
        for j in partWorkers:
            if i in checklist:
                if i in workabledays[j] :
                    checklist.remove(i)
                    workingdays[j].append(i)

except ValueError:
    pass

looptime = 1
while len(checklist) >= 1:
    try:
        for days in checklist:
            num = 0
            p = ""
            for full in fullWorkers:
                if days in workabledays[full]:
                    if num == 0:
                        num = len(workingdays[full])
                        p = full

                    elif num > len(workingdays[full]):
                        num = len(workingdays[full])
                        p = full

                    if days in workingdays[p]:
                        if p == fullWorkers[0]:
                            p = fullWorkers[1]
                        else:
                            p = fullWorkers[0]
                    checklist.remove(days)
                    workingdays[p].append(days)

    except ValueError:
        pass
    looptime += 1
    if looptime >=40:
        break

p = ""
num = 0
for l in range(0,len(partWorkers)):

    if num == 0:
        num = len(workingdays[partWorkers[l]])
        p = partWorkers[l]
    elif num <= len(workingdays[partWorkers[l]]):
        num = len(workingdays[partWorkers[l]]) # 長い方を入れている
        p = partWorkers[l]

for l in range(0,len(partWorkers)):
    if len(workingdays[partWorkers[l]]) == num:
        pass
    elif num - len(workingdays[partWorkers[l]]) >= 2:

        longerset = set(workingdays[p])
        fixingset = longerset
        fixingp = partWorkers[l]
        for d in fixingset:
            if num - len(workingdays[partWorkers[l]]) >= 2:
                if not d in workingdays[fixingp]:
                    if d in workabledays[fixingp]:
                        workingdays[p].remove(d)
                        workingdays[fixingp].append(d)

p = ""
num = 0
for l in range(0,len(fullWorkers)):

    if num == 0:
        num = len(workingdays[fullWorkers[l]])
        p = fullWorkers[l]
    elif num <= len(workingdays[fullWorkers[l]]):
        num = len(workingdays[fullWorkers[l]]) # pに長い方を入れている
        p = fullWorkers[l]
        # pは長く働く人の名前、numは長く働く人の勤務日数
for l in range(0,len(fullWorkers)):
    if len(workingdays[fullWorkers[l]]) == num:
        pass
    elif num - len(workingdays[fullWorkers[l]]) >= 2:

        longerset = set(workingdays[p])
        fixingset = longerset & set(satlist)
        fixingp = fullWorkers[l]

        for d in fixingset:
            if num - len(workingdays[fullWorkers[l]]) >= 2:
                if d not in workingdays[fixingp]:
                    if d in workabledays[fixingp]:
                        workingdays[p].remove(d)
                        print(f"{p}から{fixingp}へ{d}を移動")
                        workingdays[fixingp].append(d)

wb.create_sheet(index=2, title="workingday")
sh2 = wb["workingday"]
sh2.append(['出勤日'])
sh2.append(daysforexcel)
for j in range(len(workers)):
    tictac = []
    for days in daysInTheMonth:
        if days in workingdays[workers[j]]:
            tictac.append("○")
        else:
            tictac.append("×")
    workingexcel = [workers[j]] + tictac
    sh2.append(workingexcel)

satcell = PatternFill(
    patternType = 'solid',
    start_color='ff00ff00',
    end_color='ff0000ff')
suncell = PatternFill(
    patternType = 'solid',
    start_color='ffff0000',
    end_color='ffff0000')

for i in range(2,b+2):
    selected = sh2.cell(row=2, column=i)
    selected.alignment = Alignment( horizontal = 'center')
    selected.font = Font(size=15, bold=True)
    if selected.value in satlist:
        selected.fill = satcell
    elif selected.value in holydaylist:
        selected.fill = suncell

wb.save('シフト表.xlsx')

hardworker = ""
wdays = 0
for h in workers:
    if wdays == 0:
        wdays = len(workabledays[h])
        hardworker = h

    elif len(workabledays[h]) >= wdays:
        wdays = len(workabledays[h])
        hardworker = h

sheet.append(['日勤シフト']+dates)
sheet.append(daysforexcel)

satcell = PatternFill(
    patternType = 'solid',
    start_color='ff00ff00',
    end_color='ff0000ff')
suncell = PatternFill(
    patternType = 'solid',
    start_color='ffff0000',
    end_color='ffff0000')

for i in range(1,b+2):
    selected = sheet.cell(row=2, column=i)
    selected.alignment = Alignment( horizontal = 'center')
    selected.font = Font(size=15, bold=True)
    if selected.value in satlist:
        selected.fill = satcell
    elif selected.value in holydaylist:
        selected.fill = suncell


for worker in sh2['A']:
    counter = 1
    shiftlist=[]
    if worker.value == hardworker:
        row = worker.row
        for day in sh2[row]:
            if day.value == "○":
                shiftlist.append(counter)
                if counter == 1:
                    counter = 2
                else:
                    counter = 1
            elif day.value =="×":
                shiftlist.append("×")
        shiftlist = [hardworker] + shiftlist
        sheet.append(shiftlist)
        shift[worker.value] = shiftlist
        wb.save('シフト表.xlsx')

for worker in sh2['A']:
    shiftlist = []
    if worker.value != hardworker:
        if worker.value in workers:
            row = worker.row
            for day in sh2[row]:
                if day.value == "○":
                    thiscolumn = day.column.upper()
                    valuelist = []
                    marker = "A"
                    for cel in sheet[thiscolumn]:
                        valuelist.append(cel.value)
                    if valuelist[1] in holydaylist:
                        shiftlist.append("出")
                        marker = "B"
                    valuelist.pop(1)
                    if marker == "A":
                        if 1 in valuelist:
                            shiftlist.append(2)
                        elif 2 in valuelist:
                            shiftlist.append(1)
                        else:
                            shiftlist.append(1)
                elif day.value == "×":
                    shiftlist.append("×")

            shiftlist = [worker.value] + shiftlist
            shift[worker.value] = shiftlist
            sheet.append(shiftlist)
            wb.save('シフト表.xlsx')
for row in sh.rows:
    for cell in row:
        if cell.value != None:
            cell.alignment = Alignment( horizontal = 'center')
            cell.border = Border(
                outline=True,
                left=Side(style="medium", color="FF000000"),
                right=Side(style="medium", color="FF000000"),
                top=Side(style="medium", color="FF000000"),
                bottom=Side(style="medium", color="FF000000")
                )
for row in sh2.rows:
    for cell in row:
        if cell.value != None:
            cell.alignment = Alignment( horizontal = 'center')
            cell.border = Border(
                outline=True,
                left=Side(style="medium", color="FF000000"),
                right=Side(style="medium", color="FF000000"),
                top=Side(style="medium", color="FF000000"),
                bottom=Side(style="medium", color="FF000000")
                )
for row in sheet.rows:
    for cell in row:
        if cell.value != None:
            cell.alignment = Alignment( horizontal = 'center')
            cell.border = Border(
                outline=True,
                left=Side(style="medium", color="FF000000"),
                right=Side(style="medium", color="FF000000"),
                top=Side(style="medium", color="FF000000"),
                bottom=Side(style="medium", color="FF000000")
                )
noworkingcell = PatternFill(
    patternType = 'solid',
    start_color='ff615a5a',
    end_color='ff615a5a')
for days in sheet['2']:
    if days.value == noWorkingDays[nextmonth]:
        dcolumn = days.column # 縦がcolumn 横がrow
        for selected in sheet[dcolumn]:
            selected.fill = noworkingcell

wb.save('シフト表.xlsx')
